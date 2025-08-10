import os
import json
import polars as pl
import flet as ft
from datetime import datetime
from app.utils.temas import GestorTemas
from app.funciones.sesiones import SesionManager
from app.utils.historial import GestorHistorial

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus.tableofcontents import TableOfContents
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

async def exportar_reporte_completo(datos_reporte, metadata_reporte, page):
    """Exportar reporte con selección de formato (JSON, Excel, PDF)"""
    tema = GestorTemas.obtener_tema()
    
    if not datos_reporte:
        page.open(ft.SnackBar(
            content=ft.Text("No hay datos para exportar", color=tema.TEXT_COLOR),
            bgcolor=tema.ERROR_COLOR
        ))
        return

    # Variable para almacenar directorio seleccionado
    directorio_seleccionado = None
    
    # Selector de directorio
    def seleccionar_directorio(e: ft.FilePickerResultEvent):
        nonlocal directorio_seleccionado
        if e.path:
            directorio_seleccionado = e.path
            campo_ruta.value = e.path
            campo_ruta.update()
    
    selector_directorio = ft.FilePicker(on_result=seleccionar_directorio)
    page.overlay.append(selector_directorio)
    
    # Campo de ruta
    campo_ruta = ft.TextField(
        label="Directorio de destino",
        value="exports",
        width=400,
        bgcolor=tema.INPUT_BG,
        color=tema.TEXT_COLOR,
        border_color=tema.INPUT_BORDER,
        focused_border_color=tema.PRIMARY_COLOR,
        label_style=ft.TextStyle(color=tema.TEXT_SECONDARY)
    )
    
    # Dropdown para formato
    dropdown_formato = ft.Dropdown(
        label="Formato de exportación",
        options=[
            ft.dropdown.Option("json", "JSON (.json)", style=ft.TextStyle(color=tema.TEXT_COLOR)),
            ft.dropdown.Option("excel", "Excel (.xlsx)", style=ft.TextStyle(color=tema.TEXT_COLOR)),
            ft.dropdown.Option("pdf", "PDF (.pdf)", style=ft.TextStyle(color=tema.TEXT_COLOR)) if PDF_AVAILABLE else None,
        ],
        value="json",
        width=200,
        bgcolor=tema.INPUT_BG,
        color=tema.TEXT_COLOR,
        border_color=tema.INPUT_BORDER,
        focused_border_color=tema.PRIMARY_COLOR,
        label_style=ft.TextStyle(color=tema.TEXT_SECONDARY),
        text_style=ft.TextStyle(color=tema.TEXT_COLOR, size=14)
    )
    
    # Remover None si PDF no está disponible
    if not PDF_AVAILABLE:
        dropdown_formato.options = [opt for opt in dropdown_formato.options if opt is not None]
    
    async def realizar_exportacion(e):
        ruta_destino = directorio_seleccionado or campo_ruta.value or "exports"
        formato = dropdown_formato.value
        
        # Crear directorio si no existe
        os.makedirs(ruta_destino, exist_ok=True)
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        tipo_reporte = metadata_reporte.get('tipo_reporte', 'reporte')
        
        # Mostrar indicador de carga
        mensaje_cargando = ft.AlertDialog(
            title=ft.Text("Exportando", color=tema.TEXT_COLOR),
            bgcolor=tema.CARD_COLOR,
            content=ft.Container(
                content=ft.Row(
                    controls=[
                        ft.Text(f"Exportando reporte a {formato.upper()}...", color=tema.TEXT_COLOR),
                        ft.ProgressRing(width=14, height=14, stroke_width=2, color=tema.PRIMARY_COLOR)
                    ],
                    alignment=ft.MainAxisAlignment.CENTER,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    spacing=10,
                ),
                padding=ft.Padding(20, 20, 10, 20),
            ),
            modal=True,
        )
        page.open(mensaje_cargando)
        page.update()
        
        try:
            if formato == "json":
                await exportar_reporte_json(datos_reporte, metadata_reporte, ruta_destino, timestamp, tipo_reporte, page)
            elif formato == "excel":
                await exportar_reporte_excel(datos_reporte, metadata_reporte, ruta_destino, timestamp, tipo_reporte, page)
            elif formato == "pdf" and PDF_AVAILABLE:
                await exportar_reporte_pdf(datos_reporte, metadata_reporte, ruta_destino, timestamp, tipo_reporte, page)
                
            # Cerrar indicador de carga
            page.close(mensaje_cargando)
            page.close(dialogo_exportar)
            
        except Exception as ex:
            # Cerrar indicador de carga
            page.close(mensaje_cargando)
            print(f"Error en exportación: {ex}")
            page.open(ft.SnackBar(
                content=ft.Text(f"Error al exportar: {str(ex)}", color=tema.TEXT_COLOR),
                bgcolor=tema.ERROR_COLOR
            ))
    
    # Diálogo de exportación
    dialogo_exportar = ft.AlertDialog(
        title=ft.Text("Exportar Reporte", color=tema.TEXT_COLOR),
        bgcolor=tema.CARD_COLOR,
        content=ft.Container(
            content=ft.Column([
                ft.Text("Seleccione el directorio y formato para exportar", color=tema.TEXT_COLOR),
                ft.Container(height=10),
                ft.Row([
                    campo_ruta,
                    ft.IconButton(
                        ft.Icons.FOLDER_OPEN,
                        icon_color=tema.PRIMARY_COLOR,
                        on_click=lambda e: selector_directorio.get_directory_path(),
                        tooltip="Seleccionar directorio"
                    )
                ]),
                ft.Container(height=10),
                dropdown_formato,
                ft.Container(height=10),
                ft.Text(f"Se exportarán {len(datos_reporte)} registros", 
                       color=tema.TEXT_SECONDARY, size=12),
            ], spacing=5),
            width=500,
            height=250,
            padding=ft.Padding(15, 15, 15, 15)
        ),
        actions=[
            ft.ElevatedButton(
                "Exportar",
                style=ft.ButtonStyle(
                    bgcolor=tema.BUTTON_BG,
                    color=tema.BUTTON_TEXT,
                    shape=ft.RoundedRectangleBorder(radius=tema.BORDER_RADIUS)
                ),
                on_click=realizar_exportacion
            ),
            ft.TextButton(
                "Cancelar",
                style=ft.ButtonStyle(color=tema.TEXT_SECONDARY),
                on_click=lambda e: page.close(dialogo_exportar)
            ),
        ],
        modal=True,
    )
    
    page.open(dialogo_exportar)
    page.update()

async def exportar_reporte_json(datos_reporte, metadata_reporte, ruta_destino, timestamp, tipo_reporte, page):
    """Exportar reporte a archivo JSON"""
    tema = GestorTemas.obtener_tema()
    nombre_archivo = f"reporte_{tipo_reporte}_{timestamp}.json"
    ruta_completa = os.path.join(ruta_destino, nombre_archivo)
    
    # Datos de exportación con metadatos completos
    datos_exportacion = {
        "metadata": metadata_reporte,
        "datos": datos_reporte
    }
    
    # Guardar archivo JSON
    with open(ruta_completa, 'w', encoding='utf-8') as f:
        json.dump(datos_exportacion, f, indent=2, ensure_ascii=False, default=str)
    
    # Registrar actividad
    await registrar_actividad_exportacion("JSON", tipo_reporte, len(datos_reporte), nombre_archivo)
    
    # Mostrar mensaje de éxito
    page.open(ft.SnackBar(
        content=ft.Text(f"✅ Reporte exportado: {nombre_archivo}", color=tema.TEXT_COLOR),
        bgcolor=tema.SUCCESS_COLOR
    ))

async def exportar_reporte_excel(datos_reporte, metadata_reporte, ruta_destino, timestamp, tipo_reporte, page):
    """Exportar reporte a archivo Excel"""
    tema = GestorTemas.obtener_tema()
    nombre_archivo = f"reporte_{tipo_reporte}_{timestamp}.xlsx"
    ruta_completa = os.path.join(ruta_destino, nombre_archivo)
    
    # Preparar datos para Excel - detectar estructura automáticamente
    if not datos_reporte:
        raise ValueError("No hay datos para exportar")
    
    # Convertir datos según su estructura
    datos_excel = []
    
    # Analizar primer elemento para determinar estructura
    primer_elemento = datos_reporte[0]
    
    if isinstance(primer_elemento, dict):
        # Los datos ya son diccionarios
        for item in datos_reporte:
            datos_excel.append({
                k: str(v) if v is not None else "" 
                for k, v in item.items()
            })
    elif isinstance(primer_elemento, (list, tuple)):
        # Los datos son listas/tuplas - crear columnas genéricas
        for i, item in enumerate(datos_reporte):
            fila = {}
            for j, valor in enumerate(item):
                fila[f"Columna_{j+1}"] = str(valor) if valor is not None else ""
            datos_excel.append(fila)
    else:
        # Datos simples - crear una sola columna
        for i, item in enumerate(datos_reporte):
            datos_excel.append({"Dato": str(item) if item is not None else ""})
    
    # Crear DataFrame con Polars
    if datos_excel:
        df = pl.DataFrame(datos_excel)
        
        # Escribir el archivo Excel directamente
        df.write_excel(ruta_completa, worksheet="Datos")
        
        # Crear archivo adicional con metadatos usando openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(ruta_completa)
            
            # Crear hoja de metadatos
            if "Metadatos" not in wb.sheetnames:
                ws_meta = wb.create_sheet("Metadatos")
                ws_meta.append(["Campo", "Valor"])
                ws_meta.append(["Tipo de Reporte", metadata_reporte.get('nombre_reporte', '')])
                ws_meta.append(["Fecha de Generación", metadata_reporte.get('fecha_generacion', '')])
                ws_meta.append(["Fecha Inicio", metadata_reporte.get('fecha_inicio', '')])
                ws_meta.append(["Fecha Fin", metadata_reporte.get('fecha_fin', '')])
                ws_meta.append(["Usuario Filtro", metadata_reporte.get('usuario_filtro', '')])
                ws_meta.append(["Total Registros", str(metadata_reporte.get('total_registros', 0))])
                
                # Ajustar ancho de columnas
                ws_meta.column_dimensions['A'].width = 20
                ws_meta.column_dimensions['B'].width = 30
                
                wb.save(ruta_completa)
        except Exception as e:
            print(f"Advertencia: No se pudieron agregar metadatos: {e}")
    else:
        raise ValueError("No se pudieron procesar los datos para Excel")
    
    # Registrar actividad
    await registrar_actividad_exportacion("Excel", tipo_reporte, len(datos_reporte), nombre_archivo)
    
    # Mostrar mensaje de éxito
    page.open(ft.SnackBar(
        content=ft.Text(f"📊 Reporte Excel exportado: {nombre_archivo}", color=tema.TEXT_COLOR),
        bgcolor=tema.SUCCESS_COLOR
    ))

async def exportar_reporte_pdf(datos_reporte, metadata_reporte, ruta_destino, timestamp, tipo_reporte, page):
    """Exportar reporte a archivo PDF"""
    if not PDF_AVAILABLE:
        raise ImportError("ReportLab no está disponible para exportación PDF")
    
    tema = GestorTemas.obtener_tema()
    nombre_archivo = f"reporte_{tipo_reporte}_{timestamp}.pdf"
    ruta_completa = os.path.join(ruta_destino, nombre_archivo)
    
    # Crear documento PDF
    doc = SimpleDocTemplate(ruta_completa, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para el título
    titulo_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Centrado
        textColor=colors.darkblue
    )
    
    # Título del reporte
    story.append(Paragraph(f"REPORTE: {metadata_reporte.get('nombre_reporte', 'Sin título')}", titulo_style))
    story.append(Spacer(1, 20))
    
    # Información del reporte
    info_style = styles['Normal']
    story.append(Paragraph(f"<b>Fecha de Generación:</b> {metadata_reporte.get('fecha_generacion', '')}", info_style))
    story.append(Paragraph(f"<b>Período:</b> {metadata_reporte.get('fecha_inicio', '')} - {metadata_reporte.get('fecha_fin', '')}", info_style))
    story.append(Paragraph(f"<b>Usuario Filtro:</b> {metadata_reporte.get('usuario_filtro', '')}", info_style))
    story.append(Paragraph(f"<b>Total de Registros:</b> {metadata_reporte.get('total_registros', 0)}", info_style))
    story.append(Spacer(1, 20))
    
    # Preparar datos para la tabla
    if not datos_reporte:
        story.append(Paragraph("No hay datos disponibles para mostrar.", styles['Normal']))
    else:
        # Analizar estructura de datos
        primer_elemento = datos_reporte[0]
        
        if isinstance(primer_elemento, dict):
            # Usar las claves como encabezados
            headers = list(primer_elemento.keys())
            datos_tabla = [headers]
            
            # Limitar número de columnas para que quepan en la página
            max_cols = 6
            if len(headers) > max_cols:
                headers = headers[:max_cols]
                datos_tabla = [headers]
                story.append(Paragraph(f"<i>Nota: Mostrando las primeras {max_cols} columnas de {len(primer_elemento.keys())} totales.</i>", styles['Italic']))
                story.append(Spacer(1, 10))
            
            # Agregar datos (máximo 50 filas para PDF)
            max_rows = 50
            for i, item in enumerate(datos_reporte[:max_rows]):
                fila = []
                for header in headers:
                    valor = str(item.get(header, ''))
                    # Limitar longitud de texto por celda
                    if len(valor) > 25:
                        valor = valor[:22] + "..."
                    fila.append(valor)
                datos_tabla.append(fila)
            
            if len(datos_reporte) > max_rows:
                story.append(Paragraph(f"<i>Nota: Mostrando las primeras {max_rows} filas de {len(datos_reporte)} totales.</i>", styles['Italic']))
                story.append(Spacer(1, 10))
        
        elif isinstance(primer_elemento, (list, tuple)):
            # Crear encabezados genéricos
            num_cols = min(len(primer_elemento), 6)  # Máximo 6 columnas
            headers = [f"Col_{i+1}" for i in range(num_cols)]
            datos_tabla = [headers]
            
            # Agregar datos
            for item in datos_reporte[:50]:  # Máximo 50 filas
                fila = []
                for i in range(num_cols):
                    valor = str(item[i]) if i < len(item) else ""
                    if len(valor) > 25:
                        valor = valor[:22] + "..."
                    fila.append(valor)
                datos_tabla.append(fila)
        
        else:
            # Datos simples
            datos_tabla = [["Dato"]]
            for item in datos_reporte[:50]:
                valor = str(item)
                if len(valor) > 50:
                    valor = valor[:47] + "..."
                datos_tabla.append([valor])
        
        # Crear tabla
        tabla = Table(datos_tabla)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(tabla)
    
    # Pie de página
    story.append(Spacer(1, 30))
    story.append(Paragraph(f"Generado por TotalStock - {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Italic']))
    
    # Construir PDF
    doc.build(story)
    
    # Registrar actividad
    await registrar_actividad_exportacion("PDF", tipo_reporte, len(datos_reporte), nombre_archivo)
    
    # Mostrar mensaje de éxito
    page.open(ft.SnackBar(
        content=ft.Text(f"📄 Reporte PDF exportado: {nombre_archivo}", color=tema.TEXT_COLOR),
        bgcolor=tema.SUCCESS_COLOR
    ))

async def registrar_actividad_exportacion(formato, tipo_reporte, total_registros, nombre_archivo):
    """Registrar actividad de exportación en el historial"""
    try:
        gestor_historial = GestorHistorial()
        usuario_actual = SesionManager.obtener_usuario_actual()
        
        await gestor_historial.agregar_actividad(
            tipo="exportar_reporte",
            descripcion=f"Exportó reporte '{tipo_reporte}' con {total_registros} registros a {formato}: {nombre_archivo}",
            usuario=usuario_actual.get('username', 'Usuario') if usuario_actual else 'Sistema'
        )
    except Exception as e:
        print(f"Error al registrar actividad: {e}")
